import os.path
import os
import csv

from openpyxl import Workbook, load_workbook
from openpyxl.compat import range
from openpyxl.styles import Font, PatternFill


def get_workbook(file_path):
    if os.path.isfile(file_path):
        return load_workbook(filename=file_path)
    else:
        wb = Workbook()
        wb.remove_sheet(wb.active)
        return wb


def dump_config(ws, cfg):
    def iterate_throw_config(key_prefix, d, keys, values):
        for key, value in d.items():
            new_key = key if not key_prefix else (key_prefix + '.' + key)
            if isinstance(value, dict):
                iterate_throw_config(new_key, value, keys, values)
            else:
                keys.append(new_key)
                values.append(str(value))

    keys = []
    values = []

    iterate_throw_config('', cfg.init_dict, keys, values)
    if ws is not None:
        create_table(ws, ['key', 'value'], [keys, values], title='Config')
    else:
        return (keys, values)


def create_table(ws, column_names, columns, title=None, bold_first_column=True, selected_rows=None, selected_columns=None):
    if selected_rows is None:
        selected_rows = []

    if selected_columns is None:
        selected_columns = []

    bold_ft = Font(bold=True)
    fill = PatternFill(fill_type='solid',
                       start_color='FF27E85B',
                       end_color='FF27E85B')

    #prepare data
    formated_columns = []
    for column in columns:
        formated_column = []
        for value in column:
            if isinstance(value, int):
                formated_column.append("{:,}".format(value))
            elif isinstance(value, float):
                formated_column.append("%.4f" % value)
            else:
                formated_column.append(value)
        formated_columns.append(formated_column)

    if title:
        ws.append([title])
        ws.cell(row=ws.max_row, column=1).font = bold_ft

    ws.append(column_names)
    for i in range(len(column_names)):
        ws.cell(row=ws.max_row, column=i+1).font = bold_ft

    if not formated_columns:
        return

    for i in range(len(formated_columns[0])):
        ws.append([column[i] for column in formated_columns])

        if bold_first_column:
            ws.cell(row=ws.max_row, column=1).font = bold_ft

        if i in selected_rows:
            for j in range(len(formated_columns)):
                ws.cell(row=ws.max_row, column=j+1).fill = fill

        for column_ind in selected_columns:
            ws.cell(row=ws.max_row, column=column_ind+1).fill = fill

    ws.append([])


def fill_table_worksheet(ws, column_names, columns, title=None, bold_first_column=True,
                         selected_rows=None, selected_columns=None, cfg=None):
    create_table(ws, column_names, columns, title, bold_first_column,
                 selected_rows, selected_columns)

    if cfg:
        dump_config(ws, cfg)


class ExcelReport:

    def __init__(self, file_name):
        self.file_name = file_name
        self.wb = get_workbook(file_name)
        self.ws = None

    def start_new_sheet(self, sheet_name):
        self.ws = self.wb.create_sheet(sheet_name)

    def report(self, column_names, columns, title=None, bold_first_column=True,
               selected_rows=None, selected_columns=None, cfg=None):
        fill_table_worksheet(self.ws, column_names, columns, title, bold_first_column,
                             selected_rows, selected_columns, cfg)

    def report_config(self, cfg):
        dump_config(self.ws, cfg)

    def finish_sheet(self):
        pass

    def finish_document(self):
        self.wb.save(self.file_name)


class CSVReport:

    def __init__(self, dir_name, uuid):
        self.dir_name = dir_name
        self.uuid = uuid
        if not os.path.exists(dir_name):
            os.makedirs(dir_name)

    def start_new_sheet(self, sheet_name):
        if self.uuid:
            fine_name = os.path.join(self.dir_name,
                sheet_name + '-' + str(self.uuid) + '.csv')
        else:
            fine_name = os.path.join(self.dir_name, sheet_name + '.csv')
        self.fd = open(fine_name, 'wb')

    def report(self, column_names, columns, cfg=None, **kwargs):
        formated_columns = []
        for column in columns:
            formated_column = []
            for value in column:
                if isinstance(value, int):
                    formated_column.append(value)
                    #formated_column.append("{:,}".format(value))
                elif isinstance(value, float):
                    formated_column.append("%.4f" % value)
                else:
                    formated_column.append(value)
            formated_columns.append(formated_column)

        writer = csv.DictWriter(self.fd, fieldnames=column_names)
        writer.writeheader()

        for i in range(len(formated_columns[0])):
            row = dict([(column_names[j], formated_columns[j][i])
                for j in range(len(formated_columns))])
            writer.writerow(row)

        self.report_config(cfg)


    def report_config(self, cfg):
        keys, values = dump_config(None, cfg)
        if self.uuid:
            conf_file = os.path.join(self.dir_name, "conf-%s.txt" % str(self.uuid))
        else:
            conf_file = os.path.join(self.dir_name, "conf.txt")
        conf_lines = [ "%s: %s" % (str(k), str(v)) for k, v in zip(keys, values)]
        with open(conf_file, 'w') as output:
            output.write("\n".join(conf_lines))


    def finish_sheet(self):
        self.fd.close()

    def finish_document(self):
        pass
